"""
Gera um arquivo .xlsx pronto pra download, a partir do modelo da Planilha
FOPAG (backend/templates/planilha_fopag_template.xlsx) com os valores do
PDF já escritos nas células certas.

Isso substitui a escrita ao vivo via Google Sheets API: em vez de o
usuário "conectar" a planilha, ele baixa um arquivo já preenchido e
substitui manualmente no Google Drive quando quiser (File > Import >
Replace spreadsheet). Mais simples, sem autenticação, sem CORS, sem
depender da Cloud Function ficar "no ar" pra sempre.
"""
from __future__ import annotations

import io
import os
import re

import openpyxl

import config
from parser import Achado, MES_ANO_RELATORIO_RE  # noqa: F401 - reexportado por conveniencia

TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "templates", "planilha_fopag_template.xlsx")

# Titulos que mencionam o mes/ano da folha e podem ficar desatualizados se o
# arquivo original foi reaproveitado de um mes anterior (ex: uma aba que
# ainda diz "MAIO/2026" num mes de julho). Corrigidos automaticamente com
# base no que o proprio PDF informa - nao inventamos a data, so' repassamos
# o que o relatorio ja' diz.
TITULOS_COM_MES = [
    (config.SHEET_CONSIGNACOES, "B1"),
    (config.SHEET_LIQUIDO, "A3"),
]
_MES_NO_TITULO_RE = re.compile(
    r"(JANEIRO|FEVEREIRO|MAR[ÇC]O|ABRIL|MAIO|JUNHO|JULHO|AGOSTO|SETEMBRO|OUTUBRO|NOVEMBRO|DEZEMBRO)"
    r"\s*(DE)?\s*/?\s*(DE)?\s*\d{4}(-\d+)?",
    re.IGNORECASE,
)


def _corrigir_titulos_de_mes(wb, mes_ano_label: str | None) -> list[str]:
    """Substitui o trecho de mes/ano dentro dos titulos conhecidos, mantendo
    o resto do texto igual. Devolve a lista de celulas corrigidas, para o
    usuario ver o que mudou."""
    if not mes_ano_label:
        return []
    corrigidos = []
    for sheet_name, cell in TITULOS_COM_MES:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        atual = ws[cell].value
        if isinstance(atual, str) and _MES_NO_TITULO_RE.search(atual):
            novo = _MES_NO_TITULO_RE.sub(mes_ano_label, atual, count=1)
            if novo != atual:
                ws[cell] = novo
                corrigidos.append(f"{sheet_name}!{cell}")
    return corrigidos


def gerar_xlsx(achados: list[Achado], mes_ano_label: str | None = None,
               apenas_alta_e_media: bool = True) -> tuple[bytes, int, list[str]]:
    """Abre o modelo, escreve cada Achado na célula certa, corrige os
    títulos de mês/ano desatualizados (se mes_ano_label for informado) e
    devolve os bytes do arquivo resultante. Fórmulas e formatação do
    modelo são preservadas - só células de valor e os títulos de mês são
    sobrescritos."""
    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=False)

    # Zera antes de preencher: evita que uma célula que o mês novo não
    # mencionou (ex: nenhuma consignação daquele banco esse mês) fique com
    # o valor de julho ainda no arquivo. So' zera o que este sistema e'
    # responsavel por preencher - o resto (rubrica-por-rubrica, IRRF, NES)
    # fica intocado.
    for sheet_name, cells in config.celulas_gerenciadas().items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for cell in cells:
            ws[cell] = 0

    aplicaveis = [a for a in achados if (not apenas_alta_e_media or a.confianca in ("alta", "media"))]

    nao_encontrados = []
    for a in aplicaveis:
        if a.sheet not in wb.sheetnames:
            nao_encontrados.append(f"{a.sheet}!{a.cell} (aba não existe no modelo)")
            continue
        ws = wb[a.sheet]
        ws[a.cell] = a.valor

    _corrigir_titulos_de_mes(wb, mes_ano_label)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read(), len(aplicaveis), nao_encontrados